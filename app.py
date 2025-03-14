import os
import tempfile
import shutil
import json
import re
import logging
import pandas as pd
from flask import Flask, request, render_template, send_file, redirect, url_for, flash
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "secret-key"

# Configure logging to output to the console
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# ---------- STEP 1: Read All Sheets ----------

def read_all_sheets(file_path):
    """
    Reads *all* sheets in the Excel file with header=None and dtype=str.
    Returns a dict {sheet_name: DataFrame}.
    """
    df_dict = pd.read_excel(
        file_path,
        sheet_name=None,  # read EVERY sheet
        header=None,
        engine="openpyxl",
        dtype=str
    )
    return df_dict

# ---------- Helper Functions for each sheet ----------

def is_hours_row(row):
    """Check if any cell in the row contains 'р.час' (case-insensitive)."""
    for cell in row:
        if cell and isinstance(cell, str) and "р.час" in cell.lower():
            return True
    return False

def convert_shift_value(val):
    """
    Converts a cell value if it was misinterpreted.
      - If the value is "0.5", returns "1/2".
      - If the value looks like a date string (with optional time) and the month-day is "02-03",
        returns "2/3".
      - Otherwise returns the original value.
    """
    if pd.isna(val):
        return None
    val_str = str(val).strip()
    if val_str.lower() == "nan" or val_str == "":
        return None
    if val_str == "0.5":
        return "1/2"
    date_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})(?:\s+\d{2}:\d{2}:\d{2})?$", val_str)
    if date_match:
        year, month, day = date_match.groups()
        if month == "02" and day == "03":
            return "2/3"
    return val_str

def extract_employee_rows(df):
    """
    Identifies and pairs:
      - A row where column 2 equals 'смени'
      - The immediately following row that contains 'р.час'
    Returns a list of dicts: { code, name, shift_row, hours_row }.
    """
    rows = df.values.tolist()
    employees = []
    i = 0
    unknown_counter = 1
    while i < len(rows):
        row = rows[i]
        if len(row) >= 3 and row[2] is not None and isinstance(row[2], str) and row[2].strip().lower() == "смени":
            first_cell = row[0]
            if first_cell is not None:
                first_cell_str = str(first_cell)
                if first_cell_str.endswith(".0"):
                    first_cell_str = first_cell_str[:-2]
            else:
                first_cell_str = ""
            if first_cell_str.isdigit():
                emp_code = first_cell_str
            else:
                emp_code = f"unknown_{unknown_counter}"
                unknown_counter += 1
            emp_name = row[1] if row[1] is not None else ""
            shift_row = row
            hours_row = None
            if i + 1 < len(rows):
                next_row = rows[i + 1]
                if is_hours_row(next_row):
                    hours_row = next_row
                    i += 2
                else:
                    i += 1
            else:
                i += 1
            employees.append({
                "code": emp_code,
                "name": emp_name,
                "shift_row": shift_row,
                "hours_row": hours_row
            })
        else:
            i += 1
    return employees

def parse_days_into_dicts(df, employees):
    """
    For each employee, produce a list of dictionaries (one per day) with keys:
      code, name, date, shift, hours, day
    Assumes day-of-week labels are in row 4 and day-of-month in row 5 (columns 3..33).
    Uses pd.isna() to handle missing values.
    """
    if df.shape[0] < 6:
        return []
    day_of_week_labels = df.iloc[4, 3:3+31].tolist()
    day_numbers = df.iloc[5, 3:3+31].tolist()
    all_records = []
    for emp in employees:
        shift_row = emp["shift_row"]
        hours_row = emp["hours_row"]
        if hours_row is None:
            continue
        for day_idx in range(31):
            col_index = 3 + day_idx
            if col_index >= len(shift_row) or col_index >= len(hours_row):
                continue
            shift_val = shift_row[col_index]
            hours_val = hours_row[col_index]
            if pd.isna(shift_val):
                shift_val = None
            if pd.isna(hours_val):
                hours_val = None
            shift_val = convert_shift_value(shift_val)
            day_label = day_of_week_labels[day_idx] if day_idx < len(day_of_week_labels) else None
            day_number = day_numbers[day_idx] if day_idx < len(day_numbers) else None
            if pd.isna(day_label):
                day_label = None
            if pd.isna(day_number):
                day_number = None
            record = {
                "code": emp["code"],
                "name": emp["name"],
                "date": day_number,
                "shift": shift_val,
                "hours": hours_val,
                "day": day_label
            }
            all_records.append(record)
    return all_records

# ---------- STEP 2: Combine All Sheets ----------

def read_and_combine_all_sheets(file_path):
    """
    Reads every sheet from the Excel file, merges them into a single list of daily records.
    """
    df_dict = read_all_sheets(file_path)
    all_employees = []
    all_daily_records = []
    for sheet_name, df in df_dict.items():
        app.logger.info(f"Processing sheet: {sheet_name}")
        employees = extract_employee_rows(df)
        all_employees.extend(employees)
        daily_records = parse_days_into_dicts(df, employees)
        all_daily_records.extend(daily_records)
    return all_employees, all_daily_records

# ---------- STEP 3: Advanced Logic for Overtime and Sunday Work ----------

def parse_hours(value):
    """
    Safely converts the hours value to a float.
    If the value is missing (NaN or empty string), returns 0.0.
    """
    if pd.isna(value):
        app.logger.debug(f"parse_hours: Value is missing, returning 0.0")
        return 0.0
    value_str = str(value).strip()
    if value_str.lower() == "nan" or value_str == "":
        app.logger.debug(f"parse_hours: Found '{value}' so returning 0.0")
        return 0.0
    try:
        return float(value_str)
    except (ValueError, TypeError):
        app.logger.debug(f"parse_hours: Could not convert '{value}' to float, returning 0.0")
        return 0.0

# Mapping day index to English names and unambiguous mapping:
day_index_to_name = {
    0: "Monday",
    1: "Tuesday",
    2: "Wednesday",
    3: "Thursday",
    4: "Friday",
    5: "Saturday",
    6: "Sunday"
}
unambiguous_map = {
    "В": 1,
    "Ч": 3,
    "Н": 6
}

def process_employee_entries(entries, monthly_hours):
    """
    Process a list of daily records for one employee.
    Calculates total hours, overtime (total_hours - monthly_hours),
    and computes "overtime sunday work" based on overtime portions.
    (Duty shifts are not counted in total working hours or overtime.)
    """
    entries_sorted = sorted(
        entries,
        key=lambda x: int(str(x["date"])) if x["date"] is not None and str(x["date"]).isdigit() else 0
    )
    reference_index = None
    reference_date = None
    for entry in entries_sorted:
        if entry["day"] in unambiguous_map:
            reference_index = unambiguous_map[entry["day"]]
            try:
                reference_date = int(str(entry["date"]))
            except:
                reference_date = 1
            app.logger.debug(f"Found reference: date {reference_date} with day letter {entry['day']} -> index {reference_index}")
            break
    if reference_index is None:
        reference_index = 0
        if entries_sorted and entries_sorted[0]["date"] and str(entries_sorted[0]["date"]).isdigit():
            reference_date = int(str(entries_sorted[0]["date"]))
        else:
            reference_date = 1
        app.logger.debug(f"No unambiguous day found. Defaulting to Monday for date {reference_date}")

    total_hours = 0.0
    cumulative = 0.0
    overtime_sunday_work = 0.0  # Прекувремена работа во недела
    sunday_work_hours = 0.0
    first_shift_hours = 0.0
    second_third_hours = 0.0
    third_shift_hours = 0.0
    holidays = 0.0
    dezurstva = 0.0
    hours_per_dezurstvo = 0.0

    # Define the duty shifts that should not count towards working hours/overtime
    duty_shifts = {"д", "Д", "дпр", "ДПР", "д8", "Д8", "д16", "Д16"}

    app.logger.info(f"Processing employee {entries_sorted[0]['name']} (code: {entries_sorted[0]['code']})")
    for entry in entries_sorted:
        try:
            day_val = int(str(entry["date"])) if entry["date"] is not None and str(entry["date"]).isdigit() else 0
        except:
            day_val = 0
        day_index = (reference_index + (day_val - reference_date)) % 7
        actual_day = day_index_to_name[day_index]
        shift = entry["shift"]
        hours_worked = parse_hours(entry["hours"])

        # Only count hours for overtime/total if it's NOT a duty shift.
        if shift not in duty_shifts:
            total_hours += hours_worked
            if cumulative >= monthly_hours:
                overtime_today = hours_worked
            elif cumulative + hours_worked > monthly_hours:
                overtime_today = (cumulative + hours_worked) - monthly_hours
            else:
                overtime_today = 0.0
            cumulative += hours_worked
        else:
            # For duty shifts, we do not add their hours to total working hours or cumulative overtime.
            overtime_today = 0.0

        log_message = f"Date: {day_val}, Day: {entry['day']} -> {actual_day}, Shift: {shift}, Hours: {hours_worked}"

        # Existing Sunday work adjustments remain (they are processed regardless of counting for overtime)
        if actual_day == "Sunday":
            if shift in {"24", "1/2/3"}:
                sunday_work_hours += 18
                log_message += " | Sunday shift '24'/'1/2/3' => +18h"
            elif shift == "2/3":
                sunday_work_hours += 10
                log_message += " | Sunday shift '2/3' => +10h"
            elif shift == "3":
                sunday_work_hours += 2
                log_message += " | Sunday shift '3' => +2h"
            elif shift == "1/2":
                sunday_work_hours += hours_worked
                log_message += f" | Sunday shift '1/2' => +{hours_worked}h"
            elif shift in {"д", "Д", "дпр", "ДПР", "Д16", "д16"}:
                sunday_work_hours += 18
                log_message += f" | Sunday duty shift '{shift}' => +18h"
            elif shift == "1":
                sunday_work_hours += 8
                log_message += " | Sunday shift '1' => +8h"
        elif actual_day == "Saturday" and shift in {"24", "1/2/3", "2/3", "3", "д", "Д", "дпр", "ДПР"}:
            sunday_work_hours += 6
            log_message += " | Saturday shift => +6h for Sunday"

        # Count shift-specific hours (these branches still run even for duty shifts)
        if shift == "1":
            first_shift_hours += 8
            log_message += " | shift '1' => first shift +8"
        elif shift == "1/2":
            first_shift_hours += 8
            extra = hours_worked - 8
            if extra > 0:
                second_third_hours += extra
                log_message += f" | shift '1/2' => first +8, second +{extra}"
            else:
                log_message += " | shift '1/2' => first +8"
        elif shift == "2":
            second_third_hours += 8
            log_message += " | shift '2' => second +8"
        elif shift == "2/3":
            second_third_hours += hours_worked
            third_shift_hours += 8
            log_message += " | shift '2/3' => second +8, third +8"
        elif shift == "3":
            second_third_hours += 8
            third_shift_hours += 8
            log_message += " | shift '3' => second +8, third +8"
        elif shift in {"1/2/3", "24"}:
            first_shift_hours += 8
            second_third_hours += 16
            third_shift_hours += 8
            log_message += " | shift '1/2/3'/'24' => first +8, second +8, third +8"
        elif shift in {"ГО", "го", "Го"}:
            first_shift_hours += 8
            log_message += " | shift 'ГО' (case-insensitive) => first shift +8"
        elif shift in {"СЛ", "сл", "Сл"}:
            first_shift_hours += 8
            log_message += " | shift 'СЛ' (case-insensitive) => first shift +8"
        elif shift in {"дпр", "ДПР"}:
            holidays += hours_worked
            hours_per_dezurstvo += 8
            log_message += f" | shift '{shift}' => holidays +{hours_worked}, dezurstvo +8"
        elif shift in {"д", "Д", "Д8", "д8", "Д16", "д16"}:
            dezurstva += hours_worked
            hours_per_dezurstvo += 8
            log_message += f" | shift '{shift}' => dezurstva +{hours_worked}, dezurstvo +8"

        # Overtime Sunday Work Adjustment (only for non-duty shifts since overtime_today is 0 for duty)
        if overtime_today > 0:
            if actual_day == "Sunday":
                if shift in {"24", "1/2/3", "д", "Д", "дпр", "ДПР"}:
                    overtime_sunday_work += 18
                    log_message += " | Overtime Sunday: fixed +18h"
                elif shift == "2/3":
                    overtime_sunday_work += 10
                    log_message += " | Overtime Sunday: fixed +10h"
                elif shift == "3":
                    overtime_sunday_work += 2
                    log_message += " | Overtime Sunday: fixed +2h"
                elif shift == "1/2":
                    overtime_sunday_work += overtime_today
                    log_message += f" | Overtime Sunday: using overtime portion +{overtime_today}h"
                elif shift in {"1", "2"}:
                    overtime_sunday_work += 8
                    log_message += " | Overtime Sunday: fixed +8h"
            elif actual_day == "Saturday":
                if shift in {"24", "1/2/3", "д", "2/3", "3", "Д", "дпр", "ДПР"}:
                    overtime_sunday_work += 6
                    log_message += " | Overtime Saturday: fixed +6h"
                elif shift == "2/3":
                    overtime_sunday_work += 6
                    log_message += " | Overtime Saturday: fixed +6h"
                elif shift == "3":
                    overtime_sunday_work += 2
                    log_message += " | Overtime Saturday: fixed +2h"
                elif shift == "1/2":
                    overtime_sunday_work += overtime_today
                    log_message += f" | Overtime Saturday: using overtime portion +{overtime_today}h"
                elif shift in {"1", "2"}:
                    overtime_sunday_work += 8
                    log_message += " | Overtime Saturday: fixed +8h"

        app.logger.debug(log_message)

    overtime = total_hours - monthly_hours
    app.logger.info(f"Total hours: {total_hours}, Cumulative overtime: {overtime}, Overtime Sunday Work: {overtime_sunday_work}")
    return {
        "code": entries_sorted[0]["code"] if entries_sorted else "",
        "name": entries_sorted[0]["name"] if entries_sorted else "",
        "total_hours": total_hours,  # Total working hours (excluding duty shifts)
        "overtime": overtime,
        "sunday work": sunday_work_hours,
        "overtime sunday work": overtime_sunday_work,
        "first shift": first_shift_hours,
        "second+third shift": second_third_hours,
        "third shift": third_shift_hours,
        "holidays": holidays,
        "dezurstva": dezurstva,
        "hours per dezurstvo": hours_per_dezurstvo
    }

def generate_result_xlsx(results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Резултати"
    # Updated headers with "Вкупно работни часови" after the name column:
    headers = [
        "Шифра", 
        "Име и Презиме", 
        "Вкупно работни часови",  # Total Working Hours after the name
        "Прва смена", 
        "Втора и Трета смена заедно", 
        "Трета смена", 
        "Ноќна работа дежурство по час",
        "Дежурство на празник", 
        "Дежурство", 
        "Работа во Недела", 
        "Прекувремена работа во недела",
        "Прекувремена работа", 
    ]
    ws.append(headers)
    # Update column widths accordingly (12 columns in total)
    column_widths = [12, 24, 18, 12, 24, 12, 28, 18, 12, 18, 22, 18]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for r in results:
        row = [
            r["code"],
            r["name"],
            r["total_hours"],  # Total working hours now appears after the name
            r["first shift"],
            r["second+third shift"],
            r["third shift"],
            r["hours per dezurstvo"],
            r["holidays"],
            r["dezurstva"],
            r["sunday work"],
            r["overtime sunday work"],
            r["overtime"]
        ]
        ws.append(row)
    wb.save(output_path)

# ---------- STEP 6: Flask Routes ----------

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/process", methods=["POST"])
def process_file():
    monthly_hours_input = request.form.get("monthly_hours", "").strip()
    try:
        monthly_hours = float(monthly_hours_input)
    except ValueError:
        flash("You must input just numbers for the Total Monthly Hours.")
        return redirect(url_for("index"))
    
    if "file" not in request.files:
        flash("No file part in the request.")
        return redirect(url_for("index"))
    file = request.files["file"]
    if file.filename == "":
        flash("No file selected.")
        return redirect(url_for("index"))

    try:
        temp_dir = tempfile.mkdtemp(prefix="upload_")
        file_path = os.path.join(temp_dir, file.filename)
        file.save(file_path)
        app.logger.info(f"File saved to {file_path}")

        # Read all sheets and combine them
        all_employees, all_daily_records = read_and_combine_all_sheets(file_path)

        if not all_employees:
            flash("No valid employee data found in any sheet.")
            shutil.rmtree(temp_dir)
            return redirect(url_for("index"))

        # Group daily records by employee
        emp_dict = {}
        for rec in all_daily_records:
            key = (rec["code"], rec["name"])
            emp_dict.setdefault(key, []).append(rec)

        results = []
        for key, entries in emp_dict.items():
            processed = process_employee_entries(entries, monthly_hours)
            # Only include records with non-negative overtime
            if processed["overtime"] >= 0:
                results.append(processed)

        result_xlsx_path = os.path.join(temp_dir, "result.xlsx")
        generate_result_xlsx(results, result_xlsx_path)
        app.logger.info(f"Result XLSX generated at {result_xlsx_path}")

        temp_folder_name = os.path.basename(temp_dir)
        return render_template("download.html", temp_dir=temp_folder_name, filename="result.xlsx")

    except Exception as e:
        app.logger.error(f"Error during processing: {e}")
        flash(f"An error occurred during processing: {e}")
        return redirect(url_for("index"))

@app.route("/download/<temp_dir>/<filename>")
def download_file(temp_dir, filename):
    base_temp = tempfile.gettempdir()
    temp_path = os.path.join(base_temp, temp_dir)
    file_path = os.path.join(temp_path, filename)
    try:
        return send_file(file_path, as_attachment=True)
    finally:
        try:
            shutil.rmtree(temp_path)
            app.logger.info(f"Temporary folder {temp_path} deleted.")
        except Exception as e:
            app.logger.error(f"Error deleting temporary folder {temp_path}: {e}")

if __name__ == "__main__":
    app.run(debug=True)
